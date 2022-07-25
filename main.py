from datetime import datetime
import openpyxl
import random

device_SN = 'my_deviceSN-'
path = "data.xlsx"
book = openpyxl.load_workbook(path)
sheet = book.active


def generate_date():
    d = datetime.now().strftime('%d-%m-%y %H:%M')
    return d


def generate_id_random():
    x = random.randint(100, 999)
    return x


def generate_exp_id_random():
    max_col = sheet.max_row
    months = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    now = datetime.now()
    str_mth = now.strftime('%-m')
    month = months[int(str_mth) - 1]
    day = now.strftime('%-d')
    experience_id = device_SN + month + day + '-' + str(generate_id_random())
    for i in range(1, max_col + 1):
        if experience_id == sheet.cell(row=i, column=1):
            experience_id = device_SN + month + day + '-' + str(generate_id_random())
    return experience_id


def store_id_in_xlsx():
    e = generate_exp_id_random()
    date = generate_date()

    sheet.append((e, date))
    book.save('data.xlsx')
    print("storing done")


store_id_in_xlsx()
